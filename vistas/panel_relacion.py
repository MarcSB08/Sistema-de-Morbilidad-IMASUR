"""
Módulo para visualizar la relación mensual de morbilidad en formato de tabla
y exportarla al Formato ACT oficial de IMASUR en Excel.
"""

import calendar
import os
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modelos.consultas_medicos import ConsultasMedicos

# ──────────────────────────────────────────────────────────────────────────────
# Constantes de estilo (Formato ACT Excel)
# ──────────────────────────────────────────────────────────────────────────────
_AZUL_INST = "005A9E"
_AZUL_CLARO = "BDD7EE"
_GRIS_CLARO = "D3D3D3"
_VERDE_TOT = "E2EFDA"
_BLANCO = "FFFFFF"
_GRIS_ALT = "F2F2F2"
_NARANJA_HDR = "C55A11"
_ABREV_DIAS = ["L", "M", "M", "J", "V", "S", "D"]
_NOMBRES_MESES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

_thin = Side(style="thin", color="000000")
_BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_ORDEN_GRUPOS = ["Doctor", "Licenciado"]

# Etiquetas en español y colores por grupo
_LABELS_GRUPO = {
    "Doctor": ("DOCTORES", _AZUL_INST),
    "Licenciado": ("LICENCIADOS", _NARANJA_HDR),
}


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de estilo para openpyxl
# ──────────────────────────────────────────────────────────────────────────────


def _celda(
    ws,
    ref,
    valor=None,
    bold=False,
    size=10,
    color="000000",
    bg=None,
    h_align="center",
    v_align="center",
    borde=True,
    italic=False,
):
    c = ws[ref]
    if valor is not None:
        c.value = valor
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if borde:
        c.border = _BORDE
    return c


def _rellenar_fila(ws, fila, col_ini, col_fin, bg):
    for col in range(col_ini, col_fin + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = _BORDE


def _fila_vacia(ws, fila, col_fin):
    """Inserta una fila completamente en blanco como separador visual."""
    for col in range(1, col_fin + 1):
        ws.cell(row=fila, column=col).value = None
        ws.cell(row=fila, column=col).border = Border()
        ws.cell(row=fila, column=col).fill = PatternFill(fill_type=None)
    ws.row_dimensions[fila].height = 8


# ──────────────────────────────────────────────────────────────────────────────
# Helpers internos de sección
# ──────────────────────────────────────────────────────────────────────────────


def _escribir_encabezados_columna(
    ws, fila_ini, dias_en_mes, primer_dow, col_total, letra_total
):
    """Escribe las dos filas de encabezado de días (abreviatura + número)."""
    ws.merge_cells(f"A{fila_ini}:A{fila_ini + 1}")
    _celda(
        ws, f"A{fila_ini}", "ESPECIALISTA", bold=True, bg=_AZUL_CLARO, h_align="left"
    )
    ws.merge_cells(f"B{fila_ini}:B{fila_ini + 1}")
    _celda(
        ws, f"B{fila_ini}", "ESPECIALIDAD", bold=True, bg=_AZUL_CLARO, h_align="left"
    )

    for d in range(1, 32):
        col = d + 2
        letra = get_column_letter(col)
        dow = (primer_dow + d - 1) % 7
        if d <= dias_en_mes:
            _celda(
                ws,
                f"{letra}{fila_ini}",
                _ABREV_DIAS[dow],
                bold=True,
                size=9,
                bg=_AZUL_CLARO,
            )
            _celda(ws, f"{letra}{fila_ini + 1}", d, bold=True, size=9, bg=_AZUL_CLARO)
        else:
            for f in (fila_ini, fila_ini + 1):
                ws[f"{letra}{f}"].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
                ws[f"{letra}{f}"].border = _BORDE

    ws.merge_cells(f"{letra_total}{fila_ini}:{letra_total}{fila_ini + 1}")
    _celda(ws, f"{letra_total}{fila_ini}", "TOTAL", bold=True, bg=_AZUL_CLARO)

    return fila_ini + 2


def _escribir_bloque_especialidades(
    ws, datos_grupo, fila_actual, dias_en_mes, col_total, letra_total, colores_alt
):
    """
    Escribe todas las especialidades de un grupo.
    Devuelve la fila siguiente y la lista de filas de subtotales por especialidad.
    """
    filas_subtotal = []

    for especialidad, medicos in datos_grupo.items():
        ws.merge_cells(f"A{fila_actual}:{letra_total}{fila_actual}")
        _celda(
            ws,
            f"A{fila_actual}",
            especialidad.upper(),
            bold=True,
            size=10,
            color=_BLANCO,
            bg=_AZUL_INST,
            h_align="left",
        )
        fila_actual += 1
        fila_primer_medico = fila_actual

        for idx, (medico, dias) in enumerate(medicos.items()):
            bg = colores_alt[idx % 2]
            _celda(
                ws,
                f"A{fila_actual}",
                medico,
                bold=False,
                size=10,
                bg=bg,
                h_align="left",
            )
            _celda(
                ws,
                f"B{fila_actual}",
                especialidad,
                bold=False,
                size=9,
                bg=bg,
                h_align="left",
            )
            total_medico = 0
            for d in range(1, 32):
                col = d + 2
                letra = get_column_letter(col)
                ref = f"{letra}{fila_actual}"
                if d <= dias_en_mes:
                    val = dias.get(d, 0)
                    c = ws[ref]
                    c.value = val if val > 0 else None
                    c.font = Font(name="Arial", size=9)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.border = _BORDE
                    total_medico += val
                else:
                    ws[ref].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
                    ws[ref].border = _BORDE
            c_tot = ws[f"{letra_total}{fila_actual}"]
            c_tot.value = total_medico if total_medico > 0 else None
            c_tot.font = Font(name="Arial", bold=True, size=10)
            c_tot.alignment = Alignment(horizontal="center", vertical="center")
            c_tot.fill = PatternFill("solid", fgColor=_VERDE_TOT)
            c_tot.border = _BORDE
            fila_actual += 1

        _rellenar_fila(ws, fila_actual, 1, 2, _VERDE_TOT)
        fila_ultimo_medico = fila_actual - 1
        for d in range(1, 32):
            col = d + 2
            letra = get_column_letter(col)
            ref = f"{letra}{fila_actual}"
            if d <= dias_en_mes:
                ws[
                    ref
                ].value = (
                    f"=SUM({letra}{fila_primer_medico}:{letra}{fila_ultimo_medico})"
                )
                ws[ref].font = Font(name="Arial", bold=True, size=9)
                ws[ref].alignment = Alignment(horizontal="center", vertical="center")
                ws[ref].fill = PatternFill("solid", fgColor=_VERDE_TOT)
                ws[ref].border = _BORDE
            else:
                ws[ref].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
                ws[ref].border = _BORDE
        c_sub = ws[f"{letra_total}{fila_actual}"]
        c_sub.value = (
            f"=SUM({letra_total}{fila_primer_medico}:{letra_total}{fila_ultimo_medico})"
        )
        c_sub.font = Font(name="Arial", bold=True, size=10)
        c_sub.alignment = Alignment(horizontal="center", vertical="center")
        c_sub.fill = PatternFill("solid", fgColor=_VERDE_TOT)
        c_sub.border = _BORDE
        filas_subtotal.append(fila_actual)
        fila_actual += 1

    return fila_actual, filas_subtotal


def _escribir_fila_total_grupo(
    ws,
    fila_actual,
    label,
    dias_en_mes,
    col_total,
    letra_total,
    filas_subtotal,
    bg_color,
):
    """Escribe la fila de total acumulado de un grupo profesional completo."""
    ws.merge_cells(f"A{fila_actual}:B{fila_actual}")
    _celda(
        ws,
        f"A{fila_actual}",
        label,
        bold=True,
        size=11,
        bg=bg_color,
        color=_BLANCO,
        h_align="left",
    )

    for d in range(1, 32):
        col = d + 2
        letra = get_column_letter(col)
        ref = f"{letra}{fila_actual}"
        if d <= dias_en_mes and filas_subtotal:
            rangos = "+".join(f"{letra}{f}" for f in filas_subtotal)
            ws[ref].value = f"={rangos}"
            ws[ref].font = Font(name="Arial", bold=True, size=9)
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[ref].fill = PatternFill("solid", fgColor=bg_color)
            ws[ref].border = _BORDE
        else:
            ws[ref].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
            ws[ref].border = _BORDE

    if filas_subtotal:
        rangos_tot = "+".join(f"{letra_total}{f}" for f in filas_subtotal)
        c_g = ws[f"{letra_total}{fila_actual}"]
        c_g.value = f"={rangos_tot}"
        c_g.font = Font(name="Arial", bold=True, size=11)
        c_g.alignment = Alignment(horizontal="center", vertical="center")
        c_g.fill = PatternFill("solid", fgColor=bg_color)
        c_g.border = _BORDE

    return fila_actual + 1


# ──────────────────────────────────────────────────────────────────────────────
# Generación del archivo Excel (Formato ACT)
# ──────────────────────────────────────────────────────────────────────────────


def generar_excel_formato_act(datos, mes, anio, ruta_salida):
    """
    Genera un archivo .xlsx con el Formato ACT de Morbilidad de Consulta Externa.

    Parámetros
    ----------
    datos        : dict devuelto por ConsultasMedicos.obtener_relacion_mensual()
                   { grupo_profesional: { especialidad: { medico: { 1..31: int, 'total': int } } } }
    mes          : int  (1-12)
    anio         : int
    ruta_salida  : str  ruta completa del archivo a guardar
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{_NOMBRES_MESES[mes - 1]} {anio}"

    dias_en_mes = calendar.monthrange(anio, mes)[1]
    primer_dow = calendar.weekday(anio, mes, 1)
    col_total = 34
    letra_total = get_column_letter(col_total)
    colores_alt = [_BLANCO, _GRIS_ALT]

    # Extraer Ecografía para renderizarla al final en su sección separada,
    # pero manteniéndola dentro del bloque de Doctores en el cuerpo principal.
    datos_eco = {}
    for grupo, especialidades in datos.items():
        if "Ecografía" in especialidades:
            datos_eco = especialidades["Ecografía"]
            break

    # ── 1. ENCABEZADOS INSTITUCIONALES (filas 1-4) ────────────────────────────
    for fila, texto, bold in (
        (1, "TOTAL DE PACIENTES ATENDIDOS EN LA CONSULTA EXTERNA", True),
        (
            2,
            "INSTITUTO MUNICIPAL AUTÓNOMO DE COOPERACIÓN Y ATENCIÓN PARA LA SALUD DE URBANEJA (IMASUR)",
            True,
        ),
        (3, "LECHERÍA - EDO. ANZOÁTEGUI", False),
        (4, f"{_NOMBRES_MESES[mes - 1].upper()} {anio}", True),
    ):
        ws.merge_cells(f"A{fila}:{letra_total}{fila}")
        _celda(
            ws,
            f"A{fila}",
            texto,
            bold=bold,
            size=12 if fila == 1 else 11,
            color=_BLANCO if fila == 1 else "000000",
            bg=_AZUL_INST if fila == 1 else None,
        )

    # ── 2. ENCABEZADOS DE COLUMNA (filas 5-6) ─────────────────────────────────
    fila_actual = _escribir_encabezados_columna(
        ws, 5, dias_en_mes, primer_dow, col_total, letra_total
    )  # devuelve 7

    # ── 3. GRUPOS PROFESIONALES ────────────────────────────────────────────────
    todas_filas_subtotal = []

    for grupo in _ORDEN_GRUPOS:
        if grupo not in datos:
            continue

        label_grupo, color_banner = _LABELS_GRUPO[grupo]

        # Banner de grupo profesional
        ws.merge_cells(f"A{fila_actual}:{letra_total}{fila_actual}")
        _celda(
            ws,
            f"A{fila_actual}",
            f"── {label_grupo} ──",
            bold=True,
            size=11,
            color=_BLANCO,
            bg=color_banner,
            h_align="center",
        )
        ws.row_dimensions[fila_actual].height = 20
        fila_actual += 1

        fila_actual, filas_sub_grupo = _escribir_bloque_especialidades(
            ws,
            datos[grupo],
            fila_actual,
            dias_en_mes,
            col_total,
            letra_total,
            colores_alt,
        )

        fila_actual = _escribir_fila_total_grupo(
            ws,
            fila_actual,
            f"TOTAL {label_grupo}",
            dias_en_mes,
            col_total,
            letra_total,
            filas_sub_grupo,
            color_banner,
        )
        todas_filas_subtotal.extend(filas_sub_grupo)

        _fila_vacia(ws, fila_actual, col_total)
        fila_actual += 1

    # ── 4. TOTAL GENERAL ──────────────────────────────────────────────────────
    fila_totales = fila_actual
    ws.merge_cells(f"A{fila_totales}:B{fila_totales}")
    _celda(
        ws,
        f"A{fila_totales}",
        "TOTALES GENERALES",
        bold=True,
        size=11,
        bg=_AZUL_CLARO,
        h_align="left",
    )

    for d in range(1, 32):
        col = d + 2
        letra = get_column_letter(col)
        ref = f"{letra}{fila_totales}"
        if d <= dias_en_mes and todas_filas_subtotal:
            rangos = "+".join(f"{letra}{f}" for f in todas_filas_subtotal)
            ws[ref].value = f"={rangos}"
            ws[ref].font = Font(name="Arial", bold=True, size=9)
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[ref].fill = PatternFill("solid", fgColor=_AZUL_CLARO)
            ws[ref].border = _BORDE
        else:
            ws[ref].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
            ws[ref].border = _BORDE

    if todas_filas_subtotal:
        rangos_tot = "+".join(f"{letra_total}{f}" for f in todas_filas_subtotal)
        c_g = ws[f"{letra_total}{fila_totales}"]
        c_g.value = f"={rangos_tot}"
        c_g.font = Font(name="Arial", bold=True, size=11)
        c_g.alignment = Alignment(horizontal="center", vertical="center")
        c_g.fill = PatternFill("solid", fgColor=_AZUL_CLARO)
        c_g.border = _BORDE

    fila_actual = fila_totales + 1

    # ── 5. SECCIÓN ECOGRAFÍAS (siempre al final, sección propia) ──────────────
    fila_actual += 2
    fila_actual = _escribir_encabezados_columna(
        ws, fila_actual, dias_en_mes, primer_dow, col_total, letra_total
    )

    if datos_eco:
        fila_primer_eco = fila_actual
        for idx, (medico, dias) in enumerate(datos_eco.items()):
            bg = colores_alt[idx % 2]
            _celda(
                ws,
                f"A{fila_actual}",
                medico,
                bold=False,
                size=10,
                bg=bg,
                h_align="left",
            )
            _celda(
                ws,
                f"B{fila_actual}",
                "Ecografía",
                bold=False,
                size=9,
                bg=bg,
                h_align="left",
            )
            total_eco = 0
            for d in range(1, 32):
                col = d + 2
                letra = get_column_letter(col)
                ref = f"{letra}{fila_actual}"
                val = dias.get(d, 0)
                c = ws[ref]
                c.value = val if val > 0 else None
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = _BORDE
                total_eco += val
            c_te = ws[f"{letra_total}{fila_actual}"]
            c_te.value = total_eco if total_eco > 0 else None
            c_te.font = Font(name="Arial", bold=True, size=10)
            c_te.alignment = Alignment(horizontal="center", vertical="center")
            c_te.fill = PatternFill("solid", fgColor=_VERDE_TOT)
            c_te.border = _BORDE
            fila_actual += 1
        fila_ultimo_eco = fila_actual - 1
    else:
        fila_primer_eco = fila_actual
        _celda(
            ws, f"A{fila_actual}", "Yanira Urbina", bold=False, size=10, h_align="left"
        )
        _celda(ws, f"B{fila_actual}", "ECOGRAFISTA", bold=False, size=9, h_align="left")
        for d in range(1, 32):
            ws[f"{get_column_letter(d + 2)}{fila_actual}"].border = _BORDE
        ws[f"{letra_total}{fila_actual}"].value = 0
        ws[f"{letra_total}{fila_actual}"].border = _BORDE
        fila_ultimo_eco = fila_actual
        fila_actual += 1

    _celda(
        ws,
        f"A{fila_actual}",
        "TOTALES",
        bold=True,
        size=11,
        bg=_AZUL_CLARO,
        h_align="left",
    )
    _rellenar_fila(ws, fila_actual, 2, col_total, _AZUL_CLARO)
    for d in range(1, 32):
        ref = f"{get_column_letter(d + 2)}{fila_actual}"
        if d <= dias_en_mes:
            ws[
                ref
            ].value = f"=SUM({get_column_letter(d + 2)}{fila_primer_eco}:{get_column_letter(d + 2)}{fila_ultimo_eco})"
            ws[ref].font = Font(name="Arial", bold=True, size=9)
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[ref].border = _BORDE
        else:
            ws[ref].fill = PatternFill("solid", fgColor=_GRIS_CLARO)
            ws[ref].border = _BORDE
    c_tt = ws[f"{letra_total}{fila_actual}"]
    c_tt.value = f"=SUM({letra_total}{fila_primer_eco}:{letra_total}{fila_ultimo_eco})"
    c_tt.font = Font(name="Arial", bold=True, size=11)
    c_tt.alignment = Alignment(horizontal="center", vertical="center")
    c_tt.fill = PatternFill("solid", fgColor=_AZUL_CLARO)
    c_tt.border = _BORDE
    fila_actual += 2

    # ── 6. TIPOS DE ECOS ──────────────────────────────────────────────────────
    _celda(ws, f"A{fila_actual}", "TIPOS DE ECOS", bold=True, size=10, h_align="left")
    _celda(ws, f"B{fila_actual}", "TOTALES", bold=True, size=10)
    fila_actual += 1
    for tipo in (
        "RENAL",
        "TIROIDEO",
        "ABDOMINAL",
        "MAMARIO",
        "MUSCULO ESQUELÉTICO",
        "PARTES BLANDAS",
    ):
        _celda(ws, f"A{fila_actual}", tipo, bold=False, size=10, h_align="left")
        ws[f"B{fila_actual}"].border = _BORDE
        fila_actual += 1
    fila_actual += 1

    # ── 7. PIE DE PÁGINA ──────────────────────────────────────────────────────
    for texto in (
        "FUENTE: Morbilidad de Consulta Externa",
        "ELABORADO POR: Departamento de Registro y Estadísticas de Salud",
    ):
        ws.merge_cells(f"A{fila_actual}:{get_column_letter(17)}{fila_actual}")
        _celda(
            ws,
            f"A{fila_actual}",
            texto,
            bold=False,
            size=9,
            italic=True,
            h_align="left",
            borde=False,
        )
        fila_actual += 1

    # ── 8. DIMENSIONES Y CONFIGURACIÓN ────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    for d in range(1, 32):
        ws.column_dimensions[get_column_letter(d + 2)].width = 4.5
    ws.column_dimensions[letra_total].width = 8
    for r in range(1, 7):
        ws.row_dimensions[r].height = 18
    ws.freeze_panes = "C7"
    ws.print_title_rows = "1:6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    wb.save(ruta_salida)


# ──────────────────────────────────────────────────────────────────────────────
# Panel UI
# ──────────────────────────────────────────────────────────────────────────────


class PanelRelacion(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, corner_radius=10, fg_color="transparent")

        self.db = ConsultasMedicos()
        self._datos_actuales = None

        ctk.CTkLabel(
            self,
            text="Relación Mensual de Morbilidad",
            font=ctk.CTkFont(size=24, weight="bold"),
        ).pack(pady=(20, 10))

        frame_controles = ctk.CTkFrame(self)
        frame_controles.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(frame_controles, text="Mes:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.cmb_mes = ctk.CTkComboBox(
            frame_controles, values=_NOMBRES_MESES, state="readonly"
        )
        self.cmb_mes.grid(row=0, column=1, padx=10, pady=10)
        self.cmb_mes.set("Enero")

        ctk.CTkLabel(frame_controles, text="Año:").grid(
            row=0, column=2, padx=10, pady=10
        )
        self.cmb_anio = ctk.CTkComboBox(
            frame_controles,
            values=[str(i) for i in range(1996, 2100)],
            state="readonly",
            width=100,
        )
        self.cmb_anio.grid(row=0, column=3, padx=10, pady=10)
        self.cmb_anio.set("2026")

        self.btn_generar = ctk.CTkButton(
            frame_controles,
            text="Generar Tabla",
            font=ctk.CTkFont(weight="bold"),
            command=self.generar_tabla,
        )
        self.btn_generar.grid(row=0, column=4, padx=10, pady=10)

        self.btn_exportar = ctk.CTkButton(
            frame_controles,
            text="⬇  Exportar a Excel",
            font=ctk.CTkFont(weight="bold"),
            fg_color="#1D6F42",
            hover_color="#155232",
            state="disabled",
            command=self._exportar_excel,
        )
        self.btn_exportar.grid(row=0, column=5, padx=10, pady=10)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background="#1e1e1e",
            foreground="#d4d4d4",
            rowheight=30,
            fieldbackground="#1e1e1e",
            bordercolor="#444444",
            lightcolor="#444444",
            darkcolor="#444444",
            font=("Arial", 10),
        )
        style.configure(
            "Treeview.Heading",
            background="#2d2d30",
            foreground="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            bordercolor="#444444",
        )
        style.map("Treeview.Heading", background=[("active", "#3e3e42")])
        style.map(
            "Treeview",
            background=[("selected", "#094771")],
            foreground=[("selected", "white")],
        )

        frame_tabla = ctk.CTkFrame(self)
        frame_tabla.pack(pady=10, padx=20, fill="both", expand=True)

        scroll_y = ttk.Scrollbar(frame_tabla)
        scroll_y.pack(side="right", fill="y")
        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        self.tree = ttk.Treeview(
            frame_tabla,
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
        )
        self.tree.pack(side="left", fill="both", expand=True)
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)

        self.tree.tag_configure(
            "especialidad_header",
            background="#005a9e",
            foreground="white",
            font=("Arial", 11, "bold"),
        )
        self.tree.tag_configure(
            "grupo_doctor",
            background="#005a9e",
            foreground="white",
            font=("Arial", 12, "bold"),
        )
        self.tree.tag_configure(
            "grupo_licenciado",
            background="#7B3F00",
            foreground="white",
            font=("Arial", 12, "bold"),
        )
        self.tree.tag_configure("par", background="#252526")
        self.tree.tag_configure("impar", background="#1e1e1e")

        columnas = (
            ["ESPECIALISTA", "ESPECIALIDAD"]
            + [str(i) for i in range(1, 32)]
            + ["TOTAL"]
        )
        self.tree["columns"] = columnas
        self.tree.column("#0", width=0, stretch=ctk.NO)
        self.tree.heading("#0", text="", anchor=ctk.W)

        self.tree.column("ESPECIALISTA", anchor=ctk.W, width=220)
        self.tree.heading("ESPECIALISTA", text="MÉDICO ESPECIALISTA", anchor=ctk.W)
        self.tree.column("ESPECIALIDAD", anchor=ctk.W, width=180)
        self.tree.heading("ESPECIALIDAD", text="ESPECIALIDAD", anchor=ctk.W)

        for i in range(1, 32):
            self.tree.column(str(i), anchor=ctk.CENTER, width=40)
            self.tree.heading(str(i), text=str(i), anchor=ctk.CENTER)

        self.tree.column("TOTAL", anchor=ctk.CENTER, width=80)
        self.tree.heading("TOTAL", text="TOTAL", anchor=ctk.CENTER)

    def generar_tabla(self):
        mes_texto = self.cmb_mes.get()
        anio = int(self.cmb_anio.get())
        mes_num = _NOMBRES_MESES.index(mes_texto) + 1

        datos = self.db.obtener_relacion_mensual(mes_num, anio)

        for item in self.tree.get_children():
            self.tree.delete(item)

        if not datos:
            self._datos_actuales = None
            self.btn_exportar.configure(state="disabled")
            messagebox.showinfo(
                "Información",
                f"No hay registros de morbilidad para {mes_texto} {anio}.",
            )
            return

        self._datos_actuales = (datos, mes_num, anio, mes_texto)
        self.btn_exportar.configure(state="normal")

        _TAG_GRUPO = {
            "Doctor": "grupo_doctor",
            "Licenciado": "grupo_licenciado",
        }

        for grupo in _ORDEN_GRUPOS:
            if grupo not in datos:
                continue

            label_grupo, _ = _LABELS_GRUPO[grupo]
            tag_grupo = _TAG_GRUPO.get(grupo, "grupo_doctor")

            self.tree.insert(
                "",
                ctk.END,
                values=(f"── {label_grupo} ──", "", *[""] * 31, ""),
                tags=(tag_grupo,),
            )

            for especialidad, medicos in datos[grupo].items():
                # Ecografía se muestra dentro del bloque de Doctores en la tabla UI
                self.tree.insert(
                    "",
                    ctk.END,
                    values=(especialidad.upper(), "", *[""] * 31, ""),
                    tags=("especialidad_header",),
                )
                for idx, (medico, dias) in enumerate(medicos.items()):
                    tag = "par" if idx % 2 == 0 else "impar"
                    valores = [medico, especialidad]
                    for i in range(1, 32):
                        valores.append(dias[i] if dias[i] > 0 else "-")
                    valores.append(dias["total"])
                    self.tree.insert("", ctk.END, values=valores, tags=(tag,))

            self.tree.insert(
                "",
                ctk.END,
                values=("", "", *[""] * 31, ""),
                tags=("impar",),
            )

    def _exportar_excel(self):
        if not self._datos_actuales:
            messagebox.showwarning(
                "Sin datos",
                "Primero genere la tabla para poder exportarla.",
            )
            return

        datos, mes_num, anio, mes_texto = self._datos_actuales

        nombre_sugerido = f"Relacion_Mensual_{mes_texto}_{anio}.xlsx"
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
            initialfile=nombre_sugerido,
            title="Guardar Relación Mensual como...",
        )

        if not ruta:
            return

        try:
            generar_excel_formato_act(datos, mes_num, anio, ruta)
            messagebox.showinfo(
                "Exportación Exitosa",
                f"El archivo Excel con el Formato ACT fue generado correctamente:\n\n{ruta}",
            )
        except Exception as e:
            messagebox.showerror(
                "Error de Exportación",
                f"Ocurrió un problema al generar el archivo:\n\n{e}",
            )
